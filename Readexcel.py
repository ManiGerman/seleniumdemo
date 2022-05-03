import openpyxl

path= "C:\Excel data/Data.xlsx"

workbook=openpyxl.load_workbook(path)  #COMMAND WILL LOAD THE WORKBOOK

sheet=workbook.active  #workbook.get_sheet_by_name("sheet")

rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)

for r in range(1,rows+1): #outer for loop
    for c in range(1,cols+1): #inner for loop
        print(sheet.cell(row=r,column=c).value,end="  ")

    print()