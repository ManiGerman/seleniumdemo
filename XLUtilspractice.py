import openpyxl

def getRowcount(file,sheetName)
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumncount(file,sheetname)
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)


def readdata(file,sheetName,rownum,colnum)
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,col=colnum).value


def writedata(file,sheetName,rownum,colnum,data)
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum,col=colnum).value=data
    workbook.save(file)


