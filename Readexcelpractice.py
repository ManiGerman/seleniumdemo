import openpyxl

path= "C:/Users/Manigandan/Documents/Selenium/Testing.xlsx"

workbook=openpyxl.load_workbook(path)

#workbook.get_sheet_by_name
sheet=workbook.active

rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)


# for r in range(1,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(row=r,column=c).value,end="    ")
#
#     print()